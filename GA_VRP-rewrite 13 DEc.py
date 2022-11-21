# -*- coding: utf-8 -*-
"""
Created on Tue Dec  7 17:27:34 2021

@author: 51732
"""

#from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import random
import time
m=100
prob_mut=5
nbgen=300



def lecteur():
    global wb
    wb = load_workbook("VRP_grande_inst.xlsx")
    global ws
    ws = wb["Feuil1"]
    global n
    n=0
    while  ws.cell(n+2,1).value is not None : 
        n=n+1
    global dist    
    dist=np.zeros([n,n], dtype=float)
    #lire les distances dans la matrice dist
    for i in range(0,n):
        for j in range(0,n):
            dist[i,j]=ws.cell(i+2,j+2).value
    global dem
    dem=np.zeros(n,dtype=int)
    for i in range(n):
        dem[i]=ws.cell(44,2+i).value
    global capa
    capa=ws.cell(45,2).value

def rand_two_closest(r,visited):    
    v1=0
    dref=10000
    for i in range(0,n):
        if(dist[r,i]<dref and visited[i]==False):
            dref=dist[r,i]
            v1=i
    v2=-1
    dref=10000
    for i in range(0,n):
        if(dist[r,i]<dref and visited[i]==False and i!=v1):
            dref=dist[r,i]
            v2=i
    a=random.randint(0,1)
    if v2==-1 or a==1:
        return v1
    else : 
        return v2
    
def closest_neighbour():
    route=np.zeros(n,dtype=int)
    #construire une route avec l'aide de rand_two_closest
    visited=np.full(n,False,dtype=bool)
    visited[0]=True
    route[0]=0
    for i in range(1,n):
        route[i]=rand_two_closest(route[i-1], visited)
        visited[route[i]]=True
    return route

def calculdistance(route):
    distance=0
    load=0
    for i in range(1,n):
        if load+dem[route[i]]<=capa :
            distance+=dist[route[i-1],route[i]]
            load+=dem[route[i]]
            # print(load)
        else:
            distance+=dist[route[i-1],0]+dist[0,route[i]]
            load=dem[route[i]]
            # print('location:',i-1,'-0-',i)
            # print('route:',route[i-1],'-0-',route[i])
    distance+=dist[route[n-1],0]
    return distance

def initialize():
    global population, costs
    costs=np.zeros(m,dtype=float)
    population=np.zeros([m,n],dtype=int)
    #initialiser la population initiale
    for i in range(m):
        route=closest_neighbour()
        for j in range(n):
            population[i,j]=route[j]
        costs[i]=calculdistance(route)
    return population,costs
def mutation1(solution,location):
    head=random.randint(1,n-1)
    tail=random.randint(1,n-1)
    while head == tail :
        tail=random.randint(1,n-1)
    inter=solution[head]
    solution[head]=solution[tail]
    solution[tail]=inter
    costsdouble[location]=calculdistance(solution)

def mutation2(solution,location):
    debut=random.randint(1,n-1)
    fin=random.randint(1,n-1)
    while debut == fin :
        fin=random.randint(1,n-1)
    inter=solution[debut]
    for i in range(debut,n-1):
        solution[i]=solution[i+1]
    for i in range(n-1,fin,-1):
        solution[i]=solution[i-1]
    solution[fin]=inter
    costsdouble[location]=calculdistance(solution)
    
def mutation3(solution,location):
    a=random.randint(1,n-1)#location debut
    b=random.randint(1,n-1)#location fin
    while a == b :
        b=random.randint(1,n-1)
    inter=np.zeros(n,dtype=float)
    for i in range(a,b):
        inter[i]=solution[i]        
    for i in range(0,abs(b-a)-1,-1):
        solution[a+i]=inter[b-i]
    costsdouble[location]=calculdistance(solution)
# def mutation(populationdouble):
#     line1=random.randint(1, 2*m-1)
#     head=random.randint(1,n-1)
#     tail=random.randint(1,n-1)
#     while head == tail :
#         tail=random.randint(1,n-1)
#     inter=populationdouble[line1][head]
#     populationdouble[line1][head]=populationdouble[line1][tail]
#     populationdouble[line1][tail]=inter
#     costsdouble[line1]=calculdistance(populationdouble[line1])


#     line2=random.randint(1, 2*m-1)
#     debut=random.randint(1,n-1)
#     fin=random.randint(1,n-1)
#     while debut == fin :
#         fin=random.randint(1,n-1)
#     inter=populationdouble[line2,debut]
#     for i in range(debut,n-1):
#         populationdouble[line2,i]=populationdouble[line2,i+1]
#     for i in range(n-1,fin,-1):
#         populationdouble[line2,i]=populationdouble[line2,i-1]
#     populationdouble[line2,fin]=inter
#     costsdouble[line2]=calculdistance(populationdouble[line2])
    
#     line3=random.randint(1, m-1)    
#     a=random.randint(1,n-1)#location debut
#     b=random.randint(1,n-1)#location fin
#     while a == b :
#         b=random.randint(1,n-1)
#     inter=np.zeros(n,dtype=float)
#     for i in range(a,b):
#         inter[i]=populationdouble[line3][i]        
#     for i in range(0,abs(b-a)-1,-1):
#         populationdouble[line3][a+i]=inter[b-i]
#     costsdouble[line3]=calculdistance(populationdouble[line3])

def cross_over(p1,p2):
    place=random.randint(1, n-1)
    enfant=np.zeros([n],dtype=int)
    used=np.full([n],False,dtype=bool)
    for i in range(place):
        enfant[i]=p1[i]
        used[p1[i]]=True
    inter=0
    for i in range(place,n):
        while inter<n:
            if used[p2[inter]]==False:
                enfant[i]=p2[inter]
                used[p2[inter]]=True
                break
            else:
                inter+=1
    return enfant

def pickup(nombre,counte):
    p=np.zeros(n,dtype=int)
    costsmin=9999
    for i in range(counte):
        intex=random.randint(0, nombre-1)    
        if costsdouble[intex]<costsmin:
            costsmin=costsdouble[intex]
            popmin=intex
    for j in range(n):
        p[j]=populationdouble[popmin][j]
    return p
            
def tournoie():
    global populationdouble,costsdouble
    populationdouble=np.zeros([2*m,n],dtype=int)
    costsdouble=np.zeros([2*m],dtype=float)
    for i in range(m):
        costsdouble[i]=costs[i]
        for j in range(n):
            populationdouble[i][j]=population[i][j] 
    for i in range(m,2*m):
        p1=pickup(m,1)
        p2=pickup(m,1)
        a=cross_over(p1,p2)       
        e=cross_over(p1, a)
        for j in range(n):
            populationdouble[i][j]=e[j]
        costsdouble[i]=calculdistance(e)
    # global bestpopulation
    # costmin=9999
    # for i in range(2*m):
    #     if costsdouble[i]<costmin:
    #         bestpopulation=i
    #         costmin=costsdouble[i]
    
    # for i in range(m,2*m,2):
    #     p1=pickup(m)
    #     p2=pickup(m)
    #     inter1=cross_over(p1,p2)
    #     inter2=cross_over(p2,p1)
    #     e1=cross_over(p1, inter1)
    #     e2=cross_over(p2, inter2)        
    #     for j in range(n):
    #         populationdouble[i][j]=e1[j]
    #         populationdouble[i+1][j]=e2[j]
    #     costsdouble[i]=calculdistance(e1)
    #     costsdouble[i+1]=calculdistance(e2)         



def selection():
    for i in range(m):
        pick=pickup(2*m, 5)
        population[i]=pick
        costs[i]=calculdistance(pick)
        
        # c1=random.randint(0, 2*m-1)
        # c2=random.randint(0, 2*m-1)
    
        # if costsdouble[c1]<costsdouble[c2] : 
        #     for j in range(n):
        #         population[i][j]=populationdouble[c1][j]
        #     costs[i]=costsdouble[c1]
                    
        # else:
        #     for j in range(n):
        #         population[i][j]=populationdouble[c2][j]
        #     costs[i]=costsdouble[c2]       
        
        
        # for j in range(n):
        #     population[0][j]=populationdouble[bestpopulation][j]
        # costs[0]=costsdouble[bestpopulation]
           
        # intex=random.random()
        # probapickc1=costsdouble[c2]/(costsdouble[c1]+costsdouble[c2])
        # print(probapickc1)
        # if intex<probapickc1 :
        #     for j in range(1,n):
        #         population[i][j]=populationdouble[c1][j]
        #     costs[i]=costsdouble[c1]
                    
        # else:
        #     for j in range(n):
        #         population[i][j]=populationdouble[c2][j]
        #     costs[i]=costsdouble[c2]                

def solutionbest():
    costmin=9999    
    for i in range(m):
        if costs[i]<costmin:
            costmin=costs[i]
            popb=i
    global afterchange
    bestsolution=np.zeros(n,dtype=int)
    afterchange=np.zeros(50,dtype=int)
    for i in range(n):
        bestsolution[i]=population[popb,i]
        afterchange[i]=population[popb,i]
        
    load=0
    h=0
    for i in range(1,n):
        if load+dem[bestsolution[i]]<=capa :
            load+=dem[bestsolution[i]]
            # print(load)
        else:
            load=dem[bestsolution[i]]
            # print(0)
            # print(load)
            for j in range(49,i+h,-1):
                afterchange[j]=afterchange[j-1]
            afterchange[i+h]=0
            h+=1

            # print('location:',i-1,'-0-',i)
            # print('route:',bestsolution[i-1],'-0-',bestsolution[i])


def main():     
    lecteur()
    initialize()
    bcosts=np.zeros(nbgen,dtype=float)
    avgcosts=np.zeros(nbgen,dtype=float)
    begin = time.time()
    for gen in range(nbgen):
        tournoie()
        for i in range(2*m):
            proba=random.randint(0, 100)
            if proba<prob_mut:
                typedemutation=random.randint(0, 2)
                if typedemutation==0:
                    # print('1-p',populationdouble[i])
                    # print('1-c',costsdouble[i])
                    mutation1(populationdouble[i],i)
                    # print('2-p',populationdouble[i])
                    # print('2-c',costsdouble[i])
                    # print('2-calculer',calculdistance(populationdouble[i]))
                if typedemutation==1:
                    mutation2(populationdouble[i],i)
                if typedemutation==2:
                    mutation3(populationdouble[i],i)
        selection()            
        bcosts[gen]=min(costs)
        avgcosts[gen]=sum(costs)/len(costs)
        # print(bcosts[gen])
        # print(avgcosts[gen])
        #crossover
        #mutation
        #selection
    solutionbest()
    # print(solutionbest())
    
    end=time.time()
    elapsed=end-begin

    pd.DataFrame({'best': bcosts,'avg': avgcosts}).plot()
    pd.DataFrame({'avgcosts-bcosts':avgcosts-bcosts}).plot()
    pd.DataFrame({'bcosts':bcosts}).plot()
    print(bcosts[nbgen-1])
    print('Temps d\'exécution : {:.3f}s - temps moy \'exécution d\'une itération : {:.3f} '.format(elapsed,elapsed/nbgen))

if __name__ == '__main__':
    main()