# -*- coding: utf-8 -*-
"""
Created on Mon Sep 20 14:29:13 2021

@author: afsar
"""
from openpyxl import load_workbook
from ortools.linear_solver import pywraplp
import numpy as np
def lecteur(fname):
    global wb
    wb = load_workbook(fname, data_only=True)
    global ws
    ws = wb["Data"]
    global K, L, P, f, r, c, lo, up
    K = ws.cell(1,2).value #nb type de vehicules
    L = ws.cell(2,2).value #nb parkings
    P = ws.cell(3,2).value #nb passagers
    
    f=np.zeros((L,K),dtype=float) # cout fixe d utiliser un vehicule de type t a partir de parking j
    r=np.zeros((P,L), dtype=float)  # cout pour un passager i de se rendre jusqu au parking j
    c=np.zeros(P,dtype=float) # cout pour un passager i de ne pas faire de van sharing
    lo=np.zeros(K, dtype=float) #nombre minimal de passagers pour un vehicule de type t
    up=np.zeros(K, dtype=float) #nombre maximal de passagers pour un vehicule de type t
    
    for i in range(L):
        for j in range(K):
            f[i,j]=ws.cell(i+5,j+1).value
    
    for i in range(P):
        for j in range(L):
            r[i,j]=ws.cell(L+5+1+i,j+1).value
    
    for i in range(P):
        c[i]=ws.cell(5+L+1+P+1,i+2).value

    for i in range(K):
        lo[i]=ws.cell(5+L+1+P+1+1+1+i,1).value
        up[i]=ws.cell(5+L+1+P+1+1+1+i,2).value

    
    
def LinMod():
    #créer le solveur
    solver = pywraplp.Solver.CreateSolver('SCIP')
    
    #créer /définir les variables
      
    
    x={}
    
    for i in range(P):
        for j in range(L):
            for t in range(K):
                x[i,j,t]=solver.IntVar(0, 1, '') #variable d'affectation de passager i au parking j, type de vehicule t
  
    y={}
    for j in range(L):
            for t in range(K):
                y[j,t]=solver.IntVar(0, 1, '') # variable d'affectation au parking j type de vehicule t
    
    w={}
    for i in range(P):
        w[i]=solver.IntVar(0,1,'') # si le passager i decide de ne pas faire de van-sharing
      
    #créer la fonction objective  
    """
    obj = solver.Objective()

    for l in range(L):
        for k in range(K):
            parking+=f[l][k]*y[k][l]
                
    for p in range(P)
        nefairevan+=c[p]*w[p]

    for l in range(L):
        for p in range(P)
            for k in range(K):
    """

               
                
    obj.SetMinimization()
    
    
    
    #contraintes
     

        
        
    status = solver.Solve()
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        print('La valeur de la fonction obj : ',solver.Objective().Value())
        for j in range(L):
            for t in range(K):
                if y[j,t].solution_value()>0.99:print('Véhicule type ',t,' affecté au parking ',j)
        
        for i in range(P):
            for j in range(L):
                for t in range(K):
                    if x[i,j,t].solution_value()>0.99:print('Passager ',i,'prend le véhicule ',t,' du parking ',j )
        for i in range(P):
            if w[i].solution_value()>0.99:print('Passager ',i,' prend sa voiture')
    else : print('Modèle infaisable')
    

    
        
def main():
    lecteur("Data-VanSharing.xlsx")
    LinMod()

    
    
if __name__=="__main__": # la fonction main est lancée avec cette expression particulière
    main()