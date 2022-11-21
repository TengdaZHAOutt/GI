# -*- coding: utf-8 -*-
"""
Created on Mon Sep 20 14:29:13 2021

@author: afsar
"""
from openpyxl import load_workbook
from ortools.linear_solver import pywraplp
from math import sqrt
import numpy as np
def lecteur(fname):
    global wb
    wb = load_workbook(fname, data_only=True)
    global ws
    ws = wb["PDP1"]
    global n, dist, capa,dem
    n=ws.cell(1,2).value
    capa=ws.cell(2,2).value
    
    cox=np.zeros(n,dtype=float)
    coy=np.zeros(n,dtype=float)
          
    dist=np.zeros((n,n),dtype=float)
    dem=np.zeros(n,dtype=float)  
    global dmax
    dmax=0
    for i in range(n):
        cox[i]=ws.cell(4+i,2).value
        coy[i]=ws.cell(4+i,3).value
        dem[i]=ws.cell(4+i,4).value
    for i in range(n):
        for j in range(n):
            dist[i,j]=sqrt((cox[i]-cox[j])**2+(coy[i]-coy[j])**2)
            dmax+=dist[i,j]
    
    
    
def LinMod():
    #créer le solveur
    solver = pywraplp.Solver.CreateSolver('SCIP')
    
    #créer /définir les variables
    x={}
    
    for i in range(n):
        for j in range(n):
            x[i,j]=solver.IntVar(0, 1, '')
    t={}
    for i in range(n):
        t[i]=solver.NumVar(0,dmax,'')
    
    f={}
    for i in range(n):
        for j in range(n):
            f[i,j]=solver.NumVar(0,capa,'')
        
    
    #créer la fonction objective  
    obj = solver.Objective()
    for i in range(n):
        for j in range(n):
            if i==j : obj.SetCoefficient(x[i,j],dmax)
            else : obj.SetCoefficient(x[i,j],dist[i,j])
    obj.SetMinimization()
    
    
    
    #contraintes
    #for i in range(n): # une entrée par client
    #    solver.Add(solver.Sum([x[i,j] for j in range(n)])==1)
        
    for i in range(1,n): # une entrée par client
        contrainte=solver.RowConstraint(1,1,'')
        for j in range(n):
            if i!=j : contrainte.SetCoefficient(x[i,j],1)

            
    for i in range(n): # nb entrée égale à nb sortie
        solver.Add(solver.Sum([x[i,j] for j in range(n)]) - solver.Sum([x[j,i] for j in range(n)])==0 )
    
    #contrainte MTZ
    for i in range(n):
        for j in range(1,n):
            if i!=j:
                solver.Add(t[j]>=t[i]+dist[i,j]-dmax*(1-x[i,j]))
    
    #contrainte flot


    #contrainte de caoa

        

        
        
    status = solver.Solve()
    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        print('Coût total = ', solver.Objective().Value())
        for i in range(n):
            for j in range(n):
                if x[i,j].solution_value()>0.9 : 
                    print('De ',i,'  vers',j,' valeur ',x[i,j].solution_value(),' ',f[i,j].solution_value())
        print('Problem solved in ', solver.wall_time()/1000, 'seconds')
        print('Problem solved in %d iterations' % solver.iterations())
        print('Problem solved in %d branch-and-bound nodes' % solver.nodes())

    
        
def main():
    lecteur("PDP.xlsx")
    LinMod()
    
    
if __name__=="__main__": # la fonction main est lancée avec cette expression particulière
    main()