#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 12 16:42:59 2021

@author: afsar
"""

from openpyxl import load_workbook
from math import sqrt
import numpy as np
def lecteur(fname):
    global wb
    wb = load_workbook(fname)
    global ws
    ws = wb["Feuil1"]
    global n
    n=0
    while ws.cell(n+2,1).value is not None:
        n+=1
          
    global dist,e,l
    dist=np.zeros((n,n),dtype=float)
    coX = np.zeros(n,dtype=float)
    coY = np.zeros(n,dtype=float)
    e= np.zeros(n,dtype=float)
    l= np.zeros(n,dtype=float)
    route = np.zeros(n, dtype=int)
    #Lire les données

    for i in range(0,n):
        coX[i]=ws.cell(2+i,2).value
        coY[i]=ws.cell(2+i,3).value
        route[i]=ws.cell(2+i,7).value
        e[i]=ws.cell(2+i,4).value
        l[i]=ws.cell(2+i,5).value


    global dmax
    dmax=0
    #calculer dmax et remplir la matrice dist
    for i in range(0,n):
        for j in range(0,n):
            dist[i][j]=sqrt((coX[i]-coX[j])*(coX[i]-coX[j])+(coY[i]-coY[j])*(coY[i]-coY[j]))
            ws.cell(2+i,11+j).value=dist[i][j]






    return route

def Verificateur(times,route):
    in_time=np.zeros(n,dtype=bool)
    #remplir le tableau boolean in_time pour vérifier si oui ou non
    #les clients sont visités dans leur fenêtre de temps
    for i in range(0,n):
        if times[i] < l[route[i]] and times[i] > e[route[i]]:
            in_time[i]="True"
        else:
            in_time[i]="False"




    return in_time
    
    
def TimeWindows(route):
    times = np.zeros(n+1,dtype=float)
    times[0]=0
    #remplir le tableau times par l'heure de service 
    #si le véhicule arrive avant le début de la fenêtre de temps, il attend
    for i in range(1,n):
        a=route[i-1]
        b=route[i]
        times[i]=times[i-1]+dist[a,b]
        "times=max(e[b],times)"
        if times[i]<e[b]:
            times[i]=e[b]
        else:
            times[i]=times[i]
                
        'ws.cell(2+i,8).value=times[i]'
    return times

def main():
    route=lecteur("TSPTW.xlsx")
    times2visite=TimeWindows(route)
    alheure=Verificateur(times2visite,route)
    for i in range(n+1):
        ws.cell(i+2,8).value=times2visite[i]
        if i<n:
            if alheure[i]: ws.cell(i+2,9).value="à l'heure"
            else : ws.cell(i+2,9).value="en retard"
    
    wb.save("TSPTW-sol.xlsx")
    
if __name__=="__main__": # la fonction main est lancée avec cette expression particulière
    main()



    "488.02"