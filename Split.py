#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Sep 28 09:23:31 2021

@author: afsar
"""
from openpyxl import load_workbook
import numpy as np

def read_data(fname : str):
    global wb, ws
    wb = load_workbook(fname)
    ws = wb["Feuil1"]
    global n
    n=0
    while  ws.cell(1,n+2).value is not None : 
        n=n+1
    global dist,dem, dmax    
    dist=np.zeros([n,n])#distancier
    dem = np.zeros(n) #demande
    tgg = np.zeros(n, dtype=int) #tournée géante
    dmax=0
    for i in range(n):
        dem[i]=ws.cell(n+2,i+2).value
        tgg[i] = ws.cell(n+3,i+2).value
        for j in range(n):
            dist[i,j] = ws.cell(i+2,j+2).value
            dmax+=dist[i,j]
    global capa
    capa = ws.cell(n+4,2).value
    return tgg

def split(tg):
    V=np.full(n,dmax,dtype=float)
    P=np.zeros(n,dtype=int)
    V[0]=0
    #coder l'algorithme de split

    for i in range(n-1):        
        D=0
        cout=0
        j=i+1
        while j<=n-1 and D<=capa:
            D=D+dem[tg[j]]
            if j==i+1:
                cout=dist[0,tg[j]]+dist[tg[j],0]
            else:
                cout=cout-dist[tg[j-1],0]+dist[tg[j-1],tg[j]]+dist[tg[j],0]
            if D <= capa:
                if V[i]+cout < V[j]:
                    V[j]=V[i]+cout
                    P[j]=i
                j+=1
                



    return V[n-1], P
    
def affiche_tournees(tg,p):
    solution=[]
    deb = p[n-1]
    fin = n-1
    while fin>0:
        route=[]
        route.append(0)
        for j in range(deb+1,fin+1):
            route.append(tg[j])
        route.append(0)
        solution.append(route)
        fin=deb
        deb=p[fin]
    return solution




def main():   
    tg=read_data(fname="Distancier.xlsx")
    print(tg)
    pred = np.zeros(n,dtype=int)
    cost,pred = split(tg)
    routes=affiche_tournees(tg,pred)
    print(routes)
    print (cost)
if __name__ == '__main__':
    main()